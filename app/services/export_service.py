"""
PDF/Excel export for the Financial Reports pages (Task Backlog 2 — Epic F's
task text mentioned this but it never made it into that epic's formal
acceptance criteria). Every export function here takes data already produced
by balance_sheet_service/journal_engine_service — it never re-queries or
recomputes, so a downloaded file is always identical to what's on screen.

Two low-level renderers (_render_xlsx, _render_pdf) are shared by all three
reports; each report's export_* function only shapes its own data into the
generic (sections, summary) input those renderers expect.
"""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.models.db_models import JournalEntry

_MEDIA_TYPES = {
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}

# (heading, column headers, rows) — rows may be empty, never omitted. Numeric
# cells stay numeric (int/float) here so xlsx keeps them computable — text
# formatting (rupiah, number_format) is applied per-format at render time.
_Section = tuple[str, list[str], list[list]]
# (label, value) pairs rendered as a bold key/value block after all sections.
# value may be numeric (formatted per-format like section cells) or a plain
# string (e.g. balance status).
_Summary = list[tuple[str, object]]


def _rupiah(n) -> str:
    return "Rp " + f"{float(n or 0):,.0f}".replace(",", ".")


def _check_format(fmt: str) -> str:
    fmt = (fmt or "").lower()
    if fmt not in _MEDIA_TYPES:
        raise ValueError(f"Format ekspor '{fmt}' tidak dikenal — gunakan 'pdf' atau 'xlsx'.")
    return fmt


_INVALID_SHEET_TITLE_CHARS = str.maketrans("", "", "\\/?*[]:")


def _render_xlsx(title: str, sections: list[_Section], summary: _Summary) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title.translate(_INVALID_SHEET_TITLE_CHARS)[:31] or "Laporan"

    bold = Font(bold=True)
    row = 1
    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=14)
    row += 2

    for heading, headers, rows in sections:
        ws.cell(row=row, column=1, value=heading).font = bold
        row += 1
        for col, header in enumerate(headers, start=1):
            ws.cell(row=row, column=col, value=header).font = bold
        row += 1
        for data_row in rows:
            for col, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = "#,##0"
            row += 1
        row += 1  # blank line between sections

    if summary:
        row += 1
        for label, value in summary:
            ws.cell(row=row, column=1, value=label).font = bold
            value_cell = ws.cell(row=row, column=2, value=value)
            value_cell.font = bold
            if isinstance(value, (int, float)):
                value_cell.number_format = "#,##0"
            row += 1

    for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_pdf(title: str, sections: list[_Section], summary: _Summary) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]), Spacer(1, 0.5 * cm)]

    def _fmt_cell(value):
        return _rupiah(value) if isinstance(value, (int, float)) else str(value)

    for heading, headers, rows in sections:
        story.append(Paragraph(heading, styles["Heading3"]))
        table_data = [headers] + [[_fmt_cell(cell) for cell in data_row] for data_row in rows]
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e9ecef")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(table)
        story.append(Spacer(1, 0.5 * cm))

    if summary:
        summary_data = [[label, _fmt_cell(value)] for label, value in summary]
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        story.append(summary_table)

    doc.build(story)
    return buf.getvalue()


def _render(fmt: str, title: str, sections: list[_Section], summary: _Summary) -> tuple[bytes, str, str]:
    fmt = _check_format(fmt)
    content = _render_xlsx(title, sections, summary) if fmt == "xlsx" else _render_pdf(title, sections, summary)
    return content, _MEDIA_TYPES[fmt], fmt


def export_balance_sheet(data: dict, fmt: str) -> tuple[bytes, str, str]:
    def _account_rows(rows: list[dict]) -> list[list]:
        return [[r["kode_akun"] or "-", r["nama_akun"], r["saldo"]] for r in rows]

    sections: list[_Section] = [
        ("Aset", ["Kode", "Akun", "Saldo"], _account_rows(data["aset"])),
        ("Kewajiban", ["Kode", "Akun", "Saldo"], _account_rows(data["kewajiban"])),
        ("Ekuitas", ["Kode", "Akun", "Saldo"], _account_rows(data["ekuitas"])),
    ]
    summary = [
        ("Total Aset", data["total_aset"]),
        ("Total Kewajiban", data["total_kewajiban"]),
        ("Total Ekuitas", data["total_ekuitas"]),
        ("Total Kewajiban + Ekuitas", data["total_kewajiban_ekuitas"]),
        ("Status", "Balance" if data["is_balanced"] else "TIDAK BALANCE"),
    ]
    title = f"Neraca per {data['as_of']}"
    return _render(fmt, title, sections, summary)


def export_profit_loss(data: dict, fmt: str) -> tuple[bytes, str, str]:
    sections: list[_Section] = [
        (
            "Rincian Beban Operasional",
            ["Kode", "Akun", "Jumlah"],
            [[item["kode_akun"], item["nama_akun"], item["jumlah"]] for item in data["beban_operasional_breakdown"]],
        ),
    ]
    summary = [
        ("Total Pendapatan", data["total_pendapatan"]),
        ("HPP", data["hpp"]),
        ("Laba Kotor", data["laba_kotor"]),
        ("Total Beban Operasional", data["total_beban_operasional"]),
        ("Laba Bersih", data["laba_bersih"]),
    ]
    title = f"Laba Rugi {data['period']['start']} s/d {data['period']['end']}"
    return _render(fmt, title, sections, summary)


def export_journal(entries: list[JournalEntry], start: datetime, end: datetime, fmt: str) -> tuple[bytes, str, str]:
    headers = ["No Jurnal", "Tanggal", "Sumber Dokumen", "Debet", "Kredit", "Nominal", "Keterangan", "Status"]
    rows = [
        [
            e.no_jurnal,
            e.tanggal.isoformat(sep=" "),
            e.sumber_dokumen,
            f"{e.kode_debet} - {e.nama_akun_debet}",
            f"{e.kode_kredit} - {e.nama_akun_kredit}",
            e.nominal,
            e.keterangan,
            e.status,
        ]
        for e in entries
    ]
    sections: list[_Section] = [("Jurnal Transaksi", headers, rows)]
    title = f"Jurnal Transaksi {start.date()} s/d {end.date()}"
    return _render(fmt, title, sections, [])
