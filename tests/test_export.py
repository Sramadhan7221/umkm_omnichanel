"""
Acceptance criteria for Task Backlog 2 (Export PDF/Excel — Neraca, Laba
Rugi, Jurnal Transaksi). Every export must reuse the same service function
the on-screen report uses (balance_sheet_service.get_balance_sheet/
get_profit_loss, journal_engine_service.list_journal_entries) — never
recompute separately — so a downloaded file is always numerically identical
to what's on screen.
"""

import io
from datetime import datetime

import pytest
from openpyxl import load_workbook

from app.services.balance_sheet_service import get_balance_sheet, get_profit_loss
from app.services.chart_of_accounts_service import seed_accounts
from app.services.journal_engine_service import post_journal


def _seed(db):
    seed_accounts(db)


def _post(db, kode_debet, kode_kredit, nominal, tanggal, **kwargs):
    return post_journal(
        db, kode_debet=kode_debet, kode_kredit=kode_kredit, nominal=nominal,
        tanggal=tanggal, sumber_dokumen="test", keterangan="test", **kwargs,
    )


def _standard_scenario(db, tanggal=datetime(2026, 8, 10)):
    _post(db, "1121", "4111", 100_000, tanggal)  # revenue
    _post(db, "5212", "1121", 5_000, tanggal)     # commission expense
    _post(db, "5110", "1131", 40_000, tanggal)    # COGS


EXPORT_ENDPOINTS = [
    ("/api/financial/balance-sheet/export?as_of=2026-08-31", "balance_sheet"),
    ("/api/financial/profit-loss/export?start=2026-08-01&end=2026-08-31", "profit_loss"),
    ("/api/financial/journal/export?start=2026-08-01&end=2026-08-31", "journal"),
]


@pytest.mark.parametrize("url,_name", EXPORT_ENDPOINTS)
@pytest.mark.parametrize("fmt", ["xlsx", "pdf"])
def test_export_endpoint_returns_valid_file(client, db, url, _name, fmt):
    _seed(db)
    _standard_scenario(db)

    response = client.get(url + f"&format={fmt}")
    assert response.status_code == 200
    assert response.content

    if fmt == "xlsx":
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert response.headers["content-disposition"].endswith('.xlsx"')
        load_workbook(io.BytesIO(response.content))  # must be a valid workbook
    else:
        assert response.headers["content-type"] == "application/pdf"
        assert response.headers["content-disposition"].endswith('.pdf"')
        assert response.content.startswith(b"%PDF")


@pytest.mark.parametrize("url,_name", EXPORT_ENDPOINTS)
@pytest.mark.parametrize("fmt", ["xlsx", "pdf"])
def test_export_endpoint_succeeds_for_period_with_no_transactions(client, db, url, _name, fmt):
    _seed(db)  # no journal entries posted at all

    response = client.get(url + f"&format={fmt}")
    assert response.status_code == 200
    assert response.content
    if fmt == "pdf":
        assert response.content.startswith(b"%PDF")
    else:
        load_workbook(io.BytesIO(response.content))


@pytest.mark.parametrize("url,_name", EXPORT_ENDPOINTS)
def test_export_endpoint_rejects_unknown_format(client, db, url, _name):
    _seed(db)
    response = client.get(url + "&format=csv")
    assert response.status_code == 400


def test_balance_sheet_export_matches_on_screen_totals(client, db):
    _seed(db)
    _standard_scenario(db)
    expected = get_balance_sheet(db, datetime(2026, 8, 31, 23, 59, 59))

    response = client.get("/api/financial/balance-sheet/export?as_of=2026-08-31&format=xlsx")
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active

    values = {cell.value for row in ws.iter_rows() for cell in row if cell.value is not None}
    assert expected["total_aset"] in values
    assert expected["total_kewajiban_ekuitas"] in values


def test_profit_loss_export_matches_on_screen_totals(client, db):
    _seed(db)
    _standard_scenario(db)
    expected = get_profit_loss(db, datetime(2026, 8, 1), datetime(2026, 8, 31))

    response = client.get("/api/financial/profit-loss/export?start=2026-08-01&end=2026-08-31&format=xlsx")
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active

    values = {cell.value for row in ws.iter_rows() for cell in row if cell.value is not None}
    assert expected["laba_bersih"] in values
    assert expected["total_pendapatan"] in values
